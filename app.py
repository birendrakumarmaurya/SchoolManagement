# app.py
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, make_response,jsonify
import mysql.connector, os, time, re, io, random
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen.canvas import Canvas   # ✅ correct import
from twilio.rest import Client

app = Flask(__name__)   # ✅ only once
app.secret_key = "yoursecretkey123"   # change to something secure
#------------------------checking------------
from functools import wraps
from flask import session, redirect, url_for, flash, request

def login_required(f):
    """A decorator to protect individual routes."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # The check mirrors the @before_request logic
        if 'user_id' not in session:
            flash("You must be logged in to access this page.", "warning")
            # If the user was trying to access a specific page, save it
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

#---------for Photo upload-----------------------
UPLOAD_FOLDER = os.path.join("static", "Uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif"}
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def make_photo_filename(username, file_ext):
    """Create safe filename from username with extension"""
    safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', username.strip())
    return f"{safe_name}{file_ext}"

# ✅ MySQL configuration
db_config = {
    "host": "localhost",
    "user": "root",
    "password": "sunbeam",   # your MySQL password
    "database": "schools",
    "port": 3306
}

# ---------------- LOGIN ----------------
@app.route("/", methods=["GET", "POST"])
def login():
    error = ""
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        if username and password:
            conn = mysql.connector.connect(**db_config)
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT id, fullname, photo, password FROM users WHERE username=%s", (username,))
            user = cursor.fetchone()
            cursor.close()
            conn.close()

            if user and (check_password_hash(user["password"], password) or password == user["password"]):
                session["userid"] = user["id"]
                session["username"] = username
                session["fullname"] = user["fullname"]

                # ✅ Always prepend folder
                if user["photo"]:
                    session["photo"] = f"Uploads/{user['photo']}"
                else:
                    session["photo"] = "Images/graduation.png"

                return redirect(url_for("dashboard"))
            else:
                error = "Invalid username or password!"
        else:
            error = "Please enter username and password!"

    return render_template("login.html", error=error)

# ---------------- SIGNUP ----------------
@app.route("/signup", methods=["GET", "POST"])
def signup():
    msg = ""
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"].strip()
        fullname = request.form["fullname"].strip()
        mobile   = request.form["mobile"].strip()
        file     = request.files.get("photo")

        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # Check if username or mobile exists
        cursor.execute("SELECT id FROM users WHERE username=%s OR mobile=%s", (username, mobile))
        if cursor.fetchone():
            msg = "❌ Username or Mobile already registered!"
        else:
            hashed_pw = generate_password_hash(password)
            photo_filename = ""

            # ✅ Handle photo upload
            if file and file.filename and allowed_file(file.filename):
                ext = os.path.splitext(file.filename)[1]
                photo_filename = make_photo_filename(username, ext)
                file.save(os.path.join(app.config["UPLOAD_FOLDER"], photo_filename))

            cursor.execute(
                "INSERT INTO users (username, password, fullname, mobile, photo) VALUES (%s, %s, %s, %s, %s)",
                (username, hashed_pw, fullname, mobile, photo_filename)
            )
            conn.commit()
            msg = "✅ Signup successful. <a href='/'>Login</a>"

        cursor.close()
        conn.close()

    return render_template("signup.html", msg=msg)


# ---------------- DASHBOARD ----------------
@app.route("/dashboard")
def dashboard():
    if "username" not in session:
        return redirect(url_for("login"))

    fullname = session.get("fullname", session["username"])
    photo = session.get("photo", "Images/graduation.png")
    return render_template("dashboard.html", fullname=fullname, photo=photo)


# ---------------- LOGOUT ----------------
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ---------------- FORGOT PASSWORD (step 1) ----------------
@app.route("/forgot", methods=["GET", "POST"])
def forgot_password():
    msg = ""
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        mobile   = request.form.get("mobile", "").strip()

        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id FROM users WHERE username=%s AND mobile=%s", (username, mobile))
        user = cursor.fetchone()
        cursor.close()
        conn.close()

        if user:
            otp = random.randint(100000, 999999)
            session["otp"] = str(otp)
            session["reset_userid"] = user["id"]
            session["otp_time"] = time.time()

            # ✅ SMS OTP (Twilio)
            try:
                client = Client("TWILIO_SID", "TWILIO_AUTH_TOKEN")
                client.messages.create(
                    body=f"Your OTP is {otp}. It will expire in 45 seconds.",
                    from_="+1234567890",   # Your Twilio number
                    to=f"+91{mobile}"
                )
                flash("OTP sent successfully!", "info")
            except Exception as e:
                print("SMS error:", e)
                flash(f"Failed to send OTP (debug: {otp})", "info")

            return redirect(url_for("reset_with_otp"))
        else:
            msg = "❌ Username and Mobile do not match!"
    return render_template("forgot.html", msg=msg)


# ---------------- OTP VERIFICATION ----------------
@app.route("/reset_with_otp", methods=["GET", "POST"])
def reset_with_otp():
    msg = ""
    if request.method == "POST":
        otp_entered = request.form.get("otp", "").strip()
        otp_stored = session.get("otp")
        otp_time = session.get("otp_time", 0)

        if not otp_stored:
            msg = "❌ No OTP generated. Please try again."
        elif otp_entered == otp_stored and time.time() - otp_time <= 45:
            session.pop("otp", None)
            session.pop("otp_time", None)
            return redirect(url_for("reset_password"))
        else:
            msg = "❌ Invalid or expired OTP!"
    return render_template("reset_with_otp.html", msg=msg)


# ---------------- RESET PASSWORD (step 2) ----------------
@app.route("/reset", methods=["GET", "POST"])
def reset_password():
    if "reset_userid" not in session:
        return redirect(url_for("forgot_password"))

    msg = ""
    if request.method == "POST":
        new_password = request.form.get("password", "").strip()
        confirm      = request.form.get("confirm", "").strip()
        file         = request.files.get("photo")

        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()

        photo_filename = None
        if file and file.filename and allowed_file(file.filename):
            ext = os.path.splitext(file.filename)[1]
            photo_filename = f"user_{session['reset_userid']}_{int(time.time())}{ext}"
            file.save(os.path.join(app.config["UPLOAD_FOLDER"], photo_filename))

        # ✅ Update password if provided
        if new_password:
            if new_password != confirm:
                msg = "❌ Passwords do not match!"
                return render_template("reset.html", msg=msg)
            elif len(new_password) < 4:
                msg = "❌ Password must be at least 4 characters!"
                return render_template("reset.html", msg=msg)

            hashed_pw = generate_password_hash(new_password)
            if photo_filename:
                cursor.execute("UPDATE users SET password=%s, photo=%s WHERE id=%s",
                               (hashed_pw, photo_filename, session["reset_userid"]))
            else:
                cursor.execute("UPDATE users SET password=%s WHERE id=%s",
                               (hashed_pw, session["reset_userid"]))
        else:
            # ✅ Only update photo
            if photo_filename:
                cursor.execute("UPDATE users SET photo=%s WHERE id=%s",
                               (photo_filename, session["reset_userid"]))

        conn.commit()
        cursor.close()
        conn.close()

        session.pop("reset_userid", None)
        msg = "✅ Password/Photo updated successfully! <a href='/'>Login</a>"

    return render_template("reset.html", msg=msg)

#--------------------------------------------------

# ✅ Database connection
def get_db_connection():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="sunbeam",  # your MySQL password
        database="schools"
    )

def safe_float(val):
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

# ---------------- ROUTES SchoolDetails----------------
@app.route("/schoolDetails", methods=["GET", "POST"])
def schoolDetails():
    # ✅ handle both GET and POST
    if request.method == "POST":
        selected_session = request.form.get("session", "")
    else:
        selected_session = request.args.get("session", "")

    conn = get_db_connection()
    cursor = conn.cursor()

    if selected_session:
        cursor.execute("SELECT * FROM schoolDetails WHERE Session=%s ORDER BY SNo DESC ", (selected_session,))
    else:
        cursor.execute("SELECT * FROM schoolDetails")

    records = cursor.fetchall()

    # ✅ fetch all sessions for combo box
    cursor.execute("SELECT DISTINCT Session FROM schoolDetails ORDER BY Session DESC")
    sessions = [r[0] for r in cursor.fetchall()]

    cursor.close()
    conn.close()

    return render_template(
        "schoolDetails.html",
        records=records,
        sessions=sessions,
        selected_session=selected_session
    )

# ---------------- DELETE SINGLE ----------------
@app.route("/deleteRecord/<int:sno>", methods=["POST"])
def deleteRecord(sno):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM schoolDetails WHERE SNo=%s", (sno,))
    conn.commit()
    cursor.close()
    conn.close()
    flash("🗑 Record deleted!", "danger")
    return redirect(url_for("schoolDetails"))

# ---------------- SAVE ALL ----------------
@app.route("/updateSchoolDetails", methods=["POST"])
def updateSchoolDetails():
    text_fields = ["Session", "classes", "sections", "month"]
    num_fields = ["admitionFees", "monthlyFees", "practicalFees", "examFees", "miscellaneousCharge"]
    all_fields = text_fields + num_fields

    snos = set(k.rsplit("_", 1)[1] for k in request.form.keys() if "_" in k)

    conn = get_db_connection()
    cursor = conn.cursor()

    for sno in snos:
        vals = []
        for f in all_fields:
            raw = request.form.get(f"{f}_{sno}")
            if f in num_fields:
                vals.append(safe_float(raw))
            else:
                vals.append(raw)

        if sno.isdigit():
            cursor.execute("""
                UPDATE schoolDetails
                SET Session=%s, classes=%s, sections=%s, month=%s,
                    admitionFees=%s, monthlyFees=%s, practicalFees=%s, examFees=%s, miscellaneousCharge=%s
                WHERE SNo=%s
            """, (*vals, int(sno)))
        else:
            cursor.execute("""
                INSERT INTO schoolDetails
                  (Session, classes, sections, month,
                   admitionFees, monthlyFees, practicalFees, examFees, miscellaneousCharge)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, tuple(vals))

    conn.commit()
    cursor.close()
    conn.close()
    flash("💾 All changes saved!", "success")
    return redirect(url_for("schoolDetails"))


    
import os
from werkzeug.utils import secure_filename

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in {"jpg","jpeg","png","gif"}

def make_photo_filename(student_name, ext):
    name_clean = "_".join(student_name.strip().split())
    return f"{name_clean}{ext}"


# ------------------ Students Management ------------------
@app.route("/studentsManagement", methods=["GET"]) 
def students_management():
    search = request.args.get("search", "")
    session = request.args.get("session", "")
    class_ = request.args.get("class", "")
    section = request.args.get("section", "")
    UPLOAD_FOLDER = os.path.join("static", "uploads")
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # ---------------- Fetch Students ----------------
    query = "SELECT * FROM StudentDetails WHERE 1=1"
    params = []

    if search:
        query += " AND (StudentName LIKE %s OR RegNo LIKE %s)"
        params.extend([f"%{search}%", f"%{search}%"])
    if session:
        query += " AND Years = %s"
        params.append(session)
    if class_:
        query += " AND Class = %s"
        params.append(class_)
    if section:
        query += " AND Section = %s"
        params.append(section)
        query += " ORDER BY Years, Class, Section, RollNo"
    cursor.execute(query, tuple(params))
    students = cursor.fetchall()

    # ---------------- Fetch Dropdowns from SchoolDetails ----------------
    cursor.execute("SELECT DISTINCT session FROM SchoolDetails ORDER BY session DESC")
    sessions = [row["session"] for row in cursor.fetchall()]

    cursor.execute("SELECT DISTINCT Classes FROM SchoolDetails ORDER BY CAST(Classes AS UNSIGNED)")
    classes = [row["Classes"] for row in cursor.fetchall()]

    cursor.execute("SELECT DISTINCT Sections FROM SchoolDetails ORDER BY Sections ASC")
    sections = [row["Sections"] for row in cursor.fetchall()]

    cursor.close()
    conn.close()

    return render_template("studentsmanagement.html",
                           students=students,
                           search=search,
                           sessions=sessions,
                           classes=classes,
                           sections=sections,
                           selected_session=session,
                           selected_class=class_,
                           selected_section=section)
#---------------------------------------------------
# ------------------ Save Column Preferences ------------------
@app.route("/save_column_prefs", methods=["POST"])
def save_column_prefs():
    data = request.get_json()
    columns = data.get("columns", "")

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE StudentDetails SET Att = %s LIMIT 1", (columns,))
    conn.commit()
    cursor.close()
    conn.close()

    return jsonify(success=True)



# ------------------ Get Column Preferences ------------------
@app.route("/get_column_prefs")
def get_column_prefs():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT Att FROM StudentDetails LIMIT 1")  # or per-user/student if needed
    row = cursor.fetchone()
    cursor.close()
    conn.close()

    if row and row["Att"]:
        return jsonify(success=True, columns=row["Att"])
    return jsonify(success=True, columns="")
    
# ------------ Add students_management route) --------------

@app.route("/studentsManagement/addForm", methods=["GET"])
def add_student_form():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Fetch sessions from SchoolDetails
    cursor.execute("SELECT DISTINCT session FROM SchoolDetails ORDER BY session DESC")
    sessions = [row["session"] for row in cursor.fetchall()]

    # Fetch classes from SchoolDetails
    cursor.execute("SELECT DISTINCT Classes FROM SchoolDetails ORDER BY CAST(Classes AS UNSIGNED)")
    classes = [row["Classes"] for row in cursor.fetchall()]

    # Fetch sections from SchoolDetails
    cursor.execute("SELECT DISTINCT Sections FROM SchoolDetails ORDER BY Sections ASC")
    sections = [row["Sections"] for row in cursor.fetchall()]

    cursor.close()
    conn.close()

    return render_template(
        "student_form_page.html",
        stu=None,
        sessions=sessions,
        classes=classes,
        sections=sections
    )


@app.route("/studentsManagement/editForm/<int:student_id>", methods=["GET"])
def edit_student_form(student_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Fetch student record
    cursor.execute("SELECT * FROM StudentDetails WHERE SNo=%s", (student_id,))
    stu = cursor.fetchone()

    # Fetch dynamic dropdowns from schoolDetails
    cursor.execute("SELECT DISTINCT session FROM schoolDetails ORDER BY session DESC")
    sessions = [row["session"] for row in cursor.fetchall()]

    cursor.execute("SELECT DISTINCT Classes FROM schoolDetails ORDER BY CAST(Classes AS UNSIGNED)")
    classes = [row["Classes"] for row in cursor.fetchall()]

    cursor.execute("SELECT DISTINCT Sections FROM schoolDetails ORDER BY Sections ASC")
    sections = [row["Sections"] for row in cursor.fetchall()]

    cursor.close()
    conn.close()

    if not stu:
        flash("❌ Student not found", "danger")
        return redirect(url_for("students_management"))

    # Prepare photo URL
    if stu.get("photo"):
        stu["photo_url"] = url_for("static", filename=f"Uploads/{stu['Photo']}")
    else:
        stu["photo_url"] = url_for("static", filename="Uploads/default.png")  # default placeholder

    return render_template(
        "student_form_page.html",
        stu=stu,
        sessions=sessions,
        classes=classes,
        sections=sections
    )

# ------------------ Add Student ------------------
@app.route("/studentsManagement/add", methods=["POST"])
def add_student():
    data = request.form
    photo_file = request.files.get("Photo")
    photo_filename = None

    # Handle photo upload
    if photo_file and allowed_file(photo_file.filename):
        ext = os.path.splitext(photo_file.filename)[1].lower()
        reg_no = data.get("RegNo", "student")   # 🔹 use RegNo instead of StudentName

        # Create safe filename
        safe_reg = re.sub(r'[^a-zA-Z0-9_-]', '_', reg_no.strip()) if reg_no else "student"
        filename = f"{safe_reg}{ext}"

        # Avoid overwriting
        counter, final_filename = 1, filename
        while os.path.exists(os.path.join(app.config["UPLOAD_FOLDER"], final_filename)):
            final_filename = f"{os.path.splitext(filename)[0]}_{counter}{ext}"
            counter += 1

        # Save photo to Uploads folder
        photo_file.save(os.path.join(app.config["UPLOAD_FOLDER"], final_filename))
        photo_filename = final_filename

    # Fields in DB
    fields = ["Years","Status","RegNo","RollNo","StudentName","MotherName","FatherName","Date_Of_Birth",
              "Class","Section","Gender","Category","MothersContact","MothersOccupation","FathersContact",
              "FathersOccupation","EmergencyContact","LocalAddress","LDistrict","LCity","LState","Lpin",
              "PermanentAddress","PDistrict","PCity","PState","Ppin","Nationality","BloodGroup",
              "Religion","Cast","Aadhar","PEN","PAN","Height","Weight","AdmissionDate","MedicinePrescription",
              "IdentificationMarks","NameOfPreviousSchool","Email"]

    values = []
    for f in fields:
        new_val = data.get(f, "").strip()
        if new_val == "":
            val = None  # optional → NULL
        else:
            if f in ["RollNo","RegNo"]:
                val = int(new_val) if new_val.isdigit() else None
            else:
                val = new_val
        values.append(val)

    # Add photo as last field
    values.append(photo_filename)
    fields.append("Photo")

    # SQL build
    placeholders = ", ".join(["%s"] * len(fields))
    sql = f"INSERT INTO StudentDetails ({', '.join(fields)}) VALUES ({placeholders})"

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(sql, tuple(values))
    conn.commit()
    cursor.close()
    conn.close()

    flash("✅ Student added successfully", "success")
    return redirect(url_for("students_management"))


# ------------------ Update Student ------------------
@app.route("/studentsManagement/update/<int:student_id>", methods=["POST"])
def update_student(student_id):
    data = request.form
    photo_file = request.files.get("Photo")
    remove_photo = data.get("remove_photo")
    photo_filename = data.get("existing_photo")  # default to existing

    # Fetch existing record
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM StudentDetails WHERE SNo=%s", (student_id,))
    existing = cursor.fetchone()

    if not existing:
        flash("❌ Student not found", "danger")
        cursor.close()
        conn.close()
        return redirect(url_for("students_management"))

    # Handle remove photo
    if remove_photo and existing.get("Photo"):
        old_path = os.path.join(app.config["UPLOAD_FOLDER"], existing["Photo"])
        if os.path.exists(old_path):
            os.remove(old_path)
        photo_filename = None

    # Handle new photo (save using RegNo)
    elif photo_file and allowed_file(photo_file.filename):
        ext = os.path.splitext(photo_file.filename)[1].lower()
        reg_no = data.get("RegNo") or existing.get("RegNo") or "student"
        filename = f"{reg_no}{ext}"

        counter, final_filename = 1, filename
        while os.path.exists(os.path.join(app.config["UPLOAD_FOLDER"], final_filename)):
            final_filename = f"{reg_no}_{counter}{ext}"
            counter += 1

        photo_file.save(os.path.join(app.config["UPLOAD_FOLDER"], final_filename))
        photo_filename = final_filename

    # Fields to update
    fields = ["Years","Status","RegNo","RollNo","StudentName","MotherName","FatherName","Date_Of_Birth",
              "Class","Section","Gender","Category","MothersContact","MothersOccupation","FathersContact",
              "FathersOccupation","EmergencyContact","LocalAddress","LDistrict","LCity","LState","Lpin",
              "PermanentAddress","PDistrict","PCity","PState","Ppin","Nationality","BloodGroup",
              "Religion","Cast","Aadhar","PEN","PAN","Height","Weight","AdmissionDate","MedicinePrescription",
              "IdentificationMarks","NameOfPreviousSchool","Email"]

    set_clause, values = [], []

    for f in fields:
        new_val = data.get(f, "").strip()
        if new_val == "":
            val = existing.get(f)  # keep old value if blank
        else:
            # Convert numeric fields to int safely
            if f in ["RollNo", "RegNo"]:
                val = int(new_val) if new_val.isdigit() else None
            else:
                val = new_val
        set_clause.append(f"{f}=%s")
        values.append(val)

    # Always update photo
    set_clause.append("Photo=%s")
    values.append(photo_filename)

    sql = f"UPDATE StudentDetails SET {', '.join(set_clause)} WHERE SNo=%s"
    values.append(student_id)

    # Execute update
    update_cursor = conn.cursor()
    update_cursor.execute(sql, tuple(values))
    conn.commit()
    update_cursor.close()
    cursor.close()
    conn.close()

    flash("✏️ Student updated successfully", "info")
    return redirect(url_for("students_management"))


# ------------------ Delete Student ------------------
@app.route("/studentsManagement/delete/<int:student_id>")
def delete_student(student_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM StudentDetails WHERE SNo=%s", (student_id,))
    conn.commit()
    cursor.close()
    conn.close()
    flash("🗑️ Student deleted successfully", "danger")
    return redirect(url_for("students_management"))
#---------------- Export to Excel -----------------
@app.route("/studentsManagement/export", methods=["POST"])
def export_import():
    session_filter = request.form.get("session")
    class_filter   = request.form.get("class")
    section_filter = request.form.get("section")

    # Connect to DB
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Build query
    query = "SELECT * FROM StudentDetails WHERE 1=1"
    params = []
    if session_filter:
        query += " AND Years=%s"
        params.append(session_filter)
    if class_filter:
        query += " AND Classes=%s"
        params.append(class_filter)
    if section_filter:
        query += " AND Sections=%s"
        params.append(section_filter)

    cursor.execute(query, tuple(params))
    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    # Remove Photo column
    filtered_rows = []
    for row in rows:
        row_copy = dict(row)
        if 'Photo' in row_copy:
            del row_copy['Photo']
            filtered_rows.append(row_copy)

    # Sort rows by Class → Section → RollNo
    def sort_key(row):
        try:
            class_num = int(row.get("Class") or 0)
        except:
            class_num = 0
        section = row.get("Section") or ""
        try:
            roll = int(row.get("RollNo") or 0)
        except:
            roll = 0
        return (class_num, section, roll)

    filtered_rows.sort(key=sort_key)

    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    if filtered_rows:
        ws.append(list(filtered_rows[0].keys()))  # headers
        for row in filtered_rows:
            ws.append(list(row.values()))
    else:
        ws.append(["No records found"])

    # Filename
    parts = []
    if session_filter: parts.append(session_filter)
    if class_filter: parts.append(f"Class{class_filter}")
    if section_filter: parts.append(f"Sec{section_filter}")
    filename = "Students_" + ("_".join(parts) if parts else "All") + ".xlsx"

    # Send to browser
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

#----------------------------------------------------------
import openpyxl
from flask import request, render_template, redirect, url_for, flash, session

# -------------------- Import Data Page --------------------
@app.route("/studentsManagement/import", methods=["GET", "POST"])
def import_data():
    if request.method == "POST":
        file = request.files.get("excel_file")
        session_filter = request.form.get("session")
        class_filter = request.form.get("class")
        section_filter = request.form.get("section")

        if not file:
            flash("Please select an Excel file", "danger")
            return redirect(request.url)

        # Load Excel file
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        # Read headers
        headers = [cell.value for cell in ws[1]]

        # Read rows
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))

            # Convert RegNo, RollNo to int if possible
            for k, v in row_dict.items():
                if k in ("RegNo", "RollNo") and v is not None:
                    try:
                        row_dict[k] = int(v)
                    except:
                        row_dict[k] = v
                else:
                    row_dict[k] = str(v) if v is not None else ""
            rows.append(row_dict)

        # Apply filters
        filtered_rows = []
        for r in rows:
            if session_filter and str(r.get("Years")) != str(session_filter):
                continue
            if class_filter and str(r.get("Class")) != str(class_filter):
                continue
            if section_filter and str(r.get("Section")) != str(section_filter):
                continue
            filtered_rows.append(r)

        # Save preview in session
        session["import_preview"] = filtered_rows
        flash(f"Loaded {len(filtered_rows)} records from Excel", "success")

        return render_template(
            "import_data.html",
            preview=filtered_rows,
            sessions=get_all_sessions(),
            classes=get_all_classes(),
            sections=get_all_sections(),
            selected_session=session_filter,
            selected_class=class_filter,
            selected_section=section_filter,
        )

    # GET request
    return render_template(
        "import_data.html",
        preview=None,
        sessions=get_all_sessions(),
        classes=get_all_classes(),
        sections=get_all_sections(),
        selected_session="",
        selected_class="",
        selected_section="",
    )

def get_all_sessions():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT Session FROM schooldetails ORDER BY Session DESC")
    sessions = [row[0] for row in cursor.fetchall()]
    cursor.close()
    conn.close()
    return sessions

def get_all_classes():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT Classes FROM schooldetails ORDER BY CAST(Classes AS UNSIGNED)")
    classes = [row[0] for row in cursor.fetchall()]
    cursor.close()
    conn.close()
    return classes

def get_all_sections():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT Sections FROM schooldetails ORDER BY Sections ASC")
    sections = [row[0] for row in cursor.fetchall()]
    cursor.close()
    conn.close()
    return sections
# -------------------- Save Imported Data --------------------
@app.route("/studentsManagement/import/save", methods=["POST"])
def import_save():
    preview = session.get("import_preview", [])
    if not preview:
        flash("⚠ No data to save. Please upload Excel first.", "warning")
        return redirect(url_for("import_data"))

    conn = get_db_connection()
    cursor = conn.cursor()

    inserted, updated, skipped = 0, 0, []

    try:
        sql = """
            INSERT INTO StudentDetails
            (Years, Status, RegNo, RollNo, StudentName, MotherName, FatherName, Date_Of_Birth,
             Class, Section, Gender, Category, MothersContact, MothersOccupation, FathersContact,
             FathersOccupation, EmergencyContact, LocalAddress, LDistrict, LCity, LState, Lpin,
             PermanentAddress, PDistrict, PCity, PState, Ppin, Nationality, BloodGroup, 
             Religion, Cast, Aadhar, PEN, PAN, Height, Weight, AdmissionDate, MedicinePrescription,
             IdentificationMarks, NameOfPreviousSchool, Email)
            VALUES
            (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,  %s, %s, %s, %s, %s, %s, %s, %s,
             %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,  %s)
            ON DUPLICATE KEY UPDATE                             RollNo=VALUES(RollNo),
              StudentName=VALUES(StudentName),                  MotherName=VALUES(MotherName),
              FatherName=VALUES(FatherName),                    Date_Of_Birth=VALUES(Date_Of_Birth),
              Class=VALUES(Class),                              Section=VALUES(Section),
              Gender=VALUES(Gender),                              Category=VALUES(Category),
              MothersContact=VALUES(MothersContact),              MothersOccupation=VALUES(MothersOccupation),
              FathersContact=VALUES(FathersContact),              FathersOccupation=VALUES(FathersOccupation),
              EmergencyContact=VALUES(EmergencyContact),          LocalAddress=VALUES(LocalAddress),
              LDistrict=VALUES(LDistrict),                        LCity=VALUES(LCity),
              LState=VALUES(LState),                              Lpin=VALUES(Lpin),
              PermanentAddress=VALUES(PermanentAddress),          PDistrict=VALUES(PDistrict),
              PCity=VALUES(PCity),                                PState=VALUES(PState),
              Ppin=VALUES(Ppin),                                  Nationality=VALUES(Nationality),
              BloodGroup=VALUES(BloodGroup),                      Religion=VALUES(Religion), Cast=VALUES(Cast),
              Aadhar=VALUES(Aadhar),                              PEN=VALUES(PEN),  PAN=VALUES(PAN),
              Height=VALUES(Height),                              Weight=VALUES(Weight),  AdmissionDate=VALUES(AdmissionDate),
              MedicinePrescription=VALUES(MedicinePrescription),  IdentificationMarks=VALUES(IdentificationMarks),
              NameOfPreviousSchool=VALUES(NameOfPreviousSchool),  Email=VALUES(Email),    Status=VALUES(Status)
        """

        for i, row in enumerate(preview, start=1):
            regno = str(row.get("RegNo", "")).strip()
            years = str(row.get("Years", "")).strip()
            if not regno or not years:
                skipped.append(i)
                continue

            values = (
                years,
                str(row.get("Status") or "").strip(), regno,            str(row.get("RollNo") or "").strip(),
                str(row.get("StudentName") or "").strip(),              str(row.get("MotherName") or "").strip(),
                str(row.get("FatherName") or "").strip(),               str(row.get("Date_Of_Birth") or "").strip(),
                str(row.get("Class") or "").strip(),                    str(row.get("Section") or "").strip(),
                str(row.get("Gender") or "").strip(),                   str(row.get("Category") or "").strip(),
                str(row.get("MothersContact") or "").strip(),           str(row.get("MothersOccupation") or "").strip(),
                str(row.get("FathersContact") or "").strip(),           str(row.get("FathersOccupation") or "").strip(),
                str(row.get("EmergencyContact") or "").strip(),         str(row.get("LocalAddress") or "").strip(),
                str(row.get("LDistrict") or "").strip(),                str(row.get("LCity") or "").strip(),
                str(row.get("LState") or "").strip(),                   str(row.get("Lpin") or "").strip(),
                str(row.get("PermanentAddress") or "").strip(),         str(row.get("PDistrict") or "").strip(),
                str(row.get("PCity") or "").strip(),                    str(row.get("PState") or "").strip(),
                str(row.get("Ppin") or "").strip(),                     str(row.get("Nationality") or "").strip(),
                str(row.get("BloodGroup") or "").strip(),               str(row.get("Religion") or "").strip(), str(row.get("Cast") or "").strip(),
                str(row.get("Aadhar") or "").strip(),                   str(row.get("PEN") or "").strip(),
                str(row.get("PAN") or "").strip(),                      str(row.get("Height") or "").strip(),
                str(row.get("Weight") or "").strip(),                   str(row.get("AdmissionDate") or "").strip(),
                str(row.get("MedicinePrescription") or "").strip(),     str(row.get("IdentificationMarks") or "").strip(),
                str(row.get("NameOfPreviousSchool") or "").strip(),     str(row.get("Email") or "").strip(),
            )
            cursor.execute(sql, values)
            if cursor.rowcount == 1:
                inserted += 1
            elif cursor.rowcount == 2:  # insert+update case
                updated += 1

        conn.commit()

        msg = f"✅ Imported: {inserted} new, {updated} updated."
        if skipped:
            msg += f" Skipped rows (missing RegNo/Years): {skipped}"
        flash(msg, "success")
        session.pop("import_preview", None)
        return redirect(url_for("import_data"))

    except Exception as e:
        conn.rollback()
        flash(f"❌ Import failed: {str(e)}", "danger")
        return redirect(url_for("import_data"))

    finally:
        cursor.close()
        conn.close()
#----------- for Pring --------------------
import os
from flask import Flask, render_template, make_response, request
from io import BytesIO
from xhtml2pdf import pisa
import mysql.connector
from datetime import datetime

# Absolute path to uploads folder
PHOTO_BASE = os.path.join(os.path.dirname(__file__), "static", "Uploads")

# Define ALL_COLUMNS globally (tuple of DB field + Display Name)
ALL_COLUMNS = [
    ("SNo", "SNo"), ("Years", "Session"), ("Status", "Status"),
    ("Class", "Class"), ("Section", "Section"),
    ("RegNo", "Reg No"), ("RollNo", "Roll No"), ("Photo", "Photo"),
    ("StudentName", "Student Name"), ("FatherName", "Father Name"),
    ("MotherName", "Mother Name"), ("Date_Of_Birth", "DOB"),
    ("EmergencyContact", "Emergency Contact"), ("Aadhar", "Aadhar"),
    ("Gender", "Gender"), ("Category", "Category"),
    ("MothersContact", "Mother's Contact"),
    ("MothersOccupation", "Mother's Occupation"),
    ("FathersContact", "Father's Contact"),
    ("FathersOccupation", "Father's Occupation"),
    ("LocalAddress", "Local Address"), ("LDistrict", "Local District"),
    ("LCity", "Local City"), ("LState", "Local State"),
    ("Lpin", "Local Pin"), ("PermanentAddress", "Permanent Address"),
    ("PDistrict", "Permanent District"), ("PCity", "Permanent City"),
    ("PState", "Permanent State"), ("Ppin", "Permanent Pin"),
    ("Nationality", "Nationality"), ("BloodGroup", "Blood Group"),
    ("Religion", "Religion"), ("Cast", "Cast"), ("PEN", "PEN"),
    ("PAN", "PAN"), ("Height", "Height"), ("Weight", "Weight"),
    ("AdmissionDate", "Admission Date"),
    ("MedicinePrescription", "Medicine Prescription"),
    ("IdentificationMarks", "Identification Marks"),
    ("NameOfPreviousSchool", "Previous School"), ("Email", "Email")
]

def get_students_from_db():
    conn = mysql.connector.connect(
        host="localhost",
        user="root",
        password="sunbeam",
        database="schools",
        port=3306
    )
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM StudentDetails")
    students = cursor.fetchall()
    cursor.close()
    conn.close()
    return students

from io import BytesIO
from flask import request, render_template, make_response
from xhtml2pdf import pisa
import mysql.connector

@app.route("/print_student_details")
def print_student_details():
    session = request.args.get("session", "").strip()
    class_  = request.args.get("class", "").strip()
    section = request.args.get("section", "").strip()
    cols_param = request.args.get("cols", "")

    query = "SELECT * FROM StudentDetails WHERE 1=1"
    params = []

    # ---------------- Apply filters same as students_management ----------------
    if session and session != "All":
        query += " AND Years = %s"
        params.append(session)
    if class_ and class_ != "All":
        query += " AND Class = %s"
        params.append(class_)
    if section and section != "All":
        query += " AND Section = %s"
        params.append(section)

    # ---------------- Sorting like in management ----------------
    query += " ORDER BY Years, Class, Section, RollNo"

    conn = mysql.connector.connect(
        host="localhost",
        user="root",
        password="sunbeam",
        database="schools",
        port=3306
    )
    cursor = conn.cursor(dictionary=True)
    cursor.execute(query, tuple(params))
    students = cursor.fetchall()
    cursor.close()
    conn.close()

    # ---------------- Selected columns ----------------
    if cols_param:
        selected_indices = [int(c) for c in cols_param.split(",") if c.strip().isdigit()]
        selected_columns = [ALL_COLUMNS[i] for i in selected_indices]
    else:
        selected_columns = ALL_COLUMNS

    # ---------------- Session year for header ----------------
    session_year = session if session else (students[0].get("Years") if students else "")

    # ---------------- Render template ----------------
    rendered = render_template(
        "PrintStudentDetails.html",
        students=students,
        columns=selected_columns,
        session_year=session_year,
        photo_base=PHOTO_BASE
    )

    # ---------------- Generate PDF ----------------
    pdf_buffer = BytesIO()
    pisa_status = pisa.CreatePDF(rendered, dest=pdf_buffer)

    if pisa_status.err:
        return "Error generating PDF", 500

    pdf_buffer.seek(0)
    response = make_response(pdf_buffer.read())
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = "inline; filename=StudentDetails.pdf"
    return response


# ---------------- MAIN ----------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=80, debug=True)
